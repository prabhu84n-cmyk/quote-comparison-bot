#!/usr/bin/env python3
"""Generates the RFQ dataset (src/data/rfq.ts) and five REAL, heterogeneous
vendor response files under public/vendor-quotes/.

The vendor files are deliberately messy: different formats, different column
names, different pricing units, a foreign currency, a partial quote, an angled
phone photo of a printed rate card and a one-line email. Nothing in the app
reads this script at runtime -- the AI extracts from the produced files.
"""
import json, os, random, textwrap, zipfile, io, math
from datetime import date, timedelta

random.seed(7)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(ROOT, "public", "vendor-quotes")
os.makedirs(OUT, exist_ok=True)

# ---------------------------------------------------------------- line items
# Corrugated packaging category. 30 lines.
RAW = [
    # sku, desc, ply, spec, qty, uom, kg_per_unit(for kg-priced vendors), base price INR/unit
    ("CB-3P-305205", "RSC Shipping Carton 305x205x150mm", 3, "3-ply, 120 GSM kraft, 16 BF, brown", 24000, "piece", 0.26, 18.40),
    ("CB-3P-406305", "RSC Shipping Carton 406x305x203mm", 3, "3-ply, 140 GSM kraft, 18 BF, brown", 18000, "piece", 0.41, 26.10),
    ("CB-5P-457356", "RSC Heavy Carton 457x356x305mm", 5, "5-ply, 180 GSM kraft, 22 BF, double wall", 15000, "piece", 0.78, 47.90),
    ("CB-5P-610457", "RSC Heavy Carton 610x457x406mm", 5, "5-ply, 180 GSM kraft, 22 BF, double wall", 9000, "piece", 1.24, 74.50),
    ("CB-7P-762610", "Export Carton 762x610x508mm", 7, "7-ply, 200 GSM, 28 BF, triple wall", 3000, "piece", 2.65, 168.00),
    ("CB-3P-229152", "Small Parts Carton 229x152x102mm", 3, "3-ply, 120 GSM, 14 BF", 30000, "piece", 0.14, 11.20),
    ("CB-5P-508406", "E-comm Carton 508x406x305mm", 5, "5-ply, 160 GSM, 20 BF, white top", 12000, "piece", 0.92, 58.30),
    ("CB-3P-356254", "Book Wrap Carton 356x254x76mm", 3, "3-ply, 130 GSM, 16 BF", 16000, "piece", 0.22, 16.80),
    ("DIE-TRAY-A", "Die-cut Display Tray, 4-colour flexo", 3, "3-ply E-flute, 4-colour flexo print", 8000, "piece", 0.19, 32.40),
    ("DIE-MAIL-B", "Die-cut Mailer Box, tuck-in flap", 3, "3-ply E-flute, single colour", 20000, "piece", 0.17, 21.60),
    ("PART-6C", "Corrugated Partition, 6-cell", 2, "2-ply insert, 100 GSM", 14000, "piece", 0.11, 9.40),
    ("PART-12C", "Corrugated Partition, 12-cell", 2, "2-ply insert, 100 GSM", 9000, "piece", 0.16, 13.10),
    ("LAY-PAD-A", "Layer Pad 1200x1000mm", 2, "2-ply pad, 150 GSM", 11000, "piece", 0.48, 22.70),
    ("LAY-PAD-B", "Layer Pad 800x600mm", 2, "2-ply pad, 150 GSM", 13000, "piece", 0.24, 12.90),
    ("ROLL-3P-1M", "Corrugated Roll 1000mm x 50m", 3, "3-ply roll, 120 GSM", 900, "roll", 14.50, 1180.00),
    ("ROLL-2P-1M", "Corrugated Roll 1000mm x 75m", 2, "2-ply roll, 100 GSM", 1200, "roll", 11.20, 860.00),
    ("SHEET-5P-A", "Corrugated Sheet 1200x800mm", 5, "5-ply sheet, 180 GSM", 7000, "sheet", 0.86, 41.20),
    ("SHEET-3P-A", "Corrugated Sheet 900x600mm", 3, "3-ply sheet, 140 GSM", 10000, "sheet", 0.34, 19.60),
    ("CB-5P-PLT", "Pallet Box 1140x1140x900mm", 5, "5-ply pallet box with lid", 1200, "piece", 6.80, 640.00),
    ("SLEEVE-A", "Corrugated Pallet Sleeve, collapsible", 5, "5-ply sleeve, 4-fold", 1500, "piece", 4.10, 385.00),
    ("EDGE-PROT", "Edge Protector 50x50x3mm x 1.2m", 0, "Laminated paper angle board", 25000, "piece", 0.21, 14.30),
    ("STRAP-PP12", "PP Strapping Roll 12mm x 2000m", 0, "Machine grade, 0.6mm", 600, "roll", 8.90, 745.00),
    ("TAPE-BOPP", "BOPP Printed Tape 48mm x 65m", 0, "2-colour brand print, 45 micron", 5000, "roll", 0.32, 38.50),
    ("STRETCH-W", "Stretch Wrap Film 500mm x 23mic", 0, "Cast film, 15kg net", 800, "roll", 15.00, 1420.00),
    ("CB-3P-VENT", "Ventilated Produce Carton 400x300x150", 3, "3-ply, wax-free, vent holes", 22000, "piece", 0.30, 24.80),
    ("CB-5P-COLD", "Cold-chain Carton 450x350x250mm", 5, "5-ply, water-resistant coating", 6000, "piece", 0.95, 96.40),
    ("INS-LINER", "Insulated Liner for CB-5P-COLD", 0, "EPE foam liner, 4mm", 6000, "piece", 0.08, 34.90),
    ("CB-3P-KIT", "Flat-pack Kit Carton 500x400x200", 3, "3-ply, self-lock base", 10000, "piece", 0.44, 29.70),
    ("HONEY-PAD", "Honeycomb Pallet Pad 1100x1100", 0, "20mm honeycomb board", 2500, "piece", 1.90, 118.00),
    ("LABEL-SHIP", "Thermal Shipping Label 100x150mm", 0, "Direct thermal, 500/roll", 3000, "roll", 1.10, 212.00),
]

TODAY = date(2026, 8, 18)
ISSUE = TODAY
LINES = []
for i, (sku, desc, ply, spec, qty, uom, kg, price) in enumerate(RAW, start=1):
    LINES.append({
        "lineNo": i,
        "sku": sku,
        "category": "Packaging",
        "subCategory": "Corrugated & protective",
        "description": desc,
        "ply": ply,
        "specification": spec,
        "quantity": qty,
        "uom": uom,
        "kgPerUnit": kg,
        "deliveryLocation": "Bangalore, India",
        "requiredBy": str(ISSUE + timedelta(days=45 + (i % 4) * 7)),
        "substituteAllowed": (i % 5 != 0),
        "mandatorySpec": spec.split(",")[0],
        "notes": "",
        "_base": price,
    })

QUESTIONNAIRE = [
    ("Q1", "Do you hold a valid ISO 9001:2015 certificate?", "boolean", True, 15),
    ("Q2", "Do you hold FSC Chain-of-Custody certification for kraft paper?", "boolean", True, 10),
    ("Q3", "What is your monthly corrugation capacity in metric tonnes?", "number", 250, 15),
    ("Q4", "Do you operate an in-house testing lab (bursting strength / ECT)?", "boolean", True, 15),
    ("Q5", "How many years have you supplied corrugated packaging?", "number", 5, 10),
    ("Q6", "Can you hold two weeks of buffer stock at your plant?", "boolean", True, 10),
    ("Q7", "What is your on-time-in-full (OTIF) performance for the last 12 months (%)?", "number", 92, 15),
    ("Q8", "Do you accept 60-day payment terms?", "boolean", True, 10),
]

# ------------------------------------------------------- vendor price factors
VENDORS = {
    "kavery": {"name": "Kavery Packaging Pvt Ltd", "factor": 1.00},
    "shakti": {"name": "Shakti Corrugators Ltd", "factor": 0.955},
    "nirvana": {"name": "Nirvana Packaging Works", "factor": 1.06},
    "pacific": {"name": "Pacific Pack Global Pte Ltd", "factor": 1.02},
    "vindhya": {"name": "Vindhya Boxes & Cartons", "factor": 0.97},
}

def p(line, vkey, jitter=0.06):
    f = VENDORS[vkey]["factor"]
    j = 1 + random.uniform(-jitter, jitter)
    return round(line["_base"] * f * j, 2)

# =============================================================== 1. KAVERY XLSX
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Commercial Offer"
ws["A1"] = "KAVERY PACKAGING PVT LTD"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Plot 44, KIADB Industrial Area, Bommasandra, Bangalore 560099"
ws["A3"] = "GSTIN: 29AAFCK4821L1ZQ  |  CIN: U21029KA2009PTC050112"
ws["A4"] = "Quotation ref KPPL/Q/2026/1188 against RFQ-2026-0417 dated 18-Aug-2026"
ws["A5"] = "Contact: Sridhar Iyer, sridhar@kaverypack.in, +91 98450 22119"
ws["A6"] = "All prices EX-WORKS Bommasandra. Freight billed separately per despatch."
hdr = ["RFQ Ln", "Our Item Code", "Description as quoted", "Pack", "Qty per pack",
       "Rate (INR)", "Rate basis", "Disc %", "GST %", "MOQ", "Lead time (days)",
       "Avail qty", "Compliance", "Remarks"]
ws.append([])
ws.append(hdr)
for c in range(1, len(hdr) + 1):
    cell = ws.cell(row=8, column=c)
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="DDE7E6")

for ln in LINES:
    # Kavery quotes cartons in bundles of 25 for the small ones -> incompatible pricing unit
    bundle = 25 if ln["uom"] == "piece" and ln["_base"] < 30 else 1
    unit = "per bundle of 25" if bundle == 25 else f"per {ln['uom']}"
    rate = round(p(ln, "kavery") * bundle, 2)
    ws.append([
        ln["lineNo"], "KP-" + ln["sku"].replace("-", ""), ln["description"],
        "Bundle" if bundle == 25 else ln["uom"].title(), bundle,
        rate, unit, 2.5 if ln["quantity"] > 12000 else 0, 18,
        bundle * 200 if bundle == 25 else 500,
        12 if ln["_base"] < 50 else 18,
        ln["quantity"], "Yes", "" if bundle == 1 else "Supplied shrink-wrapped in 25s",
    ])
ws.column_dimensions["C"].width = 46
ws.column_dimensions["G"].width = 20

ws2 = wb.create_sheet("Charges & Terms")
ws2.append(["Head", "Value", "Notes"])
ws2.append(["Freight to Bangalore", 145000, "Lump sum for full order, ex-works otherwise"])
ws2.append(["Packing & palletisation", 38000, "Lump sum"])
ws2.append(["Transit insurance", "0.35% of order value", ""])
ws2.append(["Installation / configuration", "Not applicable", ""])
ws2.append(["Quote validity", "45 days from 26-Aug-2026", ""])
ws2.append(["Payment terms", "45 days from invoice", ""])
ws2.append(["Price escalation", "Linked to IPMA kraft index, reviewed quarterly", ""])
ws2.append(["Early payment discount", "1.5% if paid within 15 days", ""])
ws2.append(["Volume rebate", "1% credit note above INR 1.5 Cr annual offtake", ""])
ws2.append(["Partial delivery", "Yes, weekly despatch schedule possible", ""])
ws2.append(["Back-order policy", "Balance shipped within 7 days at no extra freight", ""])
ws2.append(["Country of origin", "India", ""])

ws3 = wb.create_sheet("Questionnaire")
ws3.append(["#", "Question", "Our answer"])
ans = ["Yes, valid till Mar-2028", "Yes, FSC-C118842", "410", "Yes, in-house BS/ECT lab",
       "17", "Yes", "94.5", "No - we can offer 45 days only"]
for (qid, q, *_), a in zip(QUESTIONNAIRE, ans):
    ws3.append([qid, q, a])
ws3.column_dimensions["B"].width = 70
wb.save(os.path.join(OUT, "kavery-packaging-quote.xlsx"))

# ============================================================== 2. SHAKTI PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rlcanvas

path = os.path.join(OUT, "shakti-corrugators-quote.pdf")
c = rlcanvas.Canvas(path, pagesize=A4)
W, H = A4

def header(c, page):
    c.setFont("Helvetica-Bold", 16)
    c.drawString(20 * mm, H - 20 * mm, "SHAKTI CORRUGATORS LIMITED")
    c.setFont("Helvetica", 8)
    c.drawString(20 * mm, H - 25 * mm, "Survey 118/2, Hosur Road, Attibele, Bangalore Rural 562107, Karnataka, India")
    c.drawString(20 * mm, H - 29 * mm, "GSTIN 29AABCS7712M1Z4  |  CIN L21015KA1994PLC016622  |  www.shakticorrugators.co.in")
    c.line(20 * mm, H - 32 * mm, W - 20 * mm, H - 32 * mm)
    c.setFont("Helvetica", 7)
    c.drawRightString(W - 20 * mm, H - 20 * mm, "Page %d" % page)

header(c, 1)
y = H - 40 * mm
c.setFont("Helvetica-Bold", 11)
c.drawString(20 * mm, y, "QUOTATION  SCL/BLR/2026/0904")
y -= 6 * mm
c.setFont("Helvetica", 9)
for t in [
    "Against: RFQ-2026-0417  -  Corrugated & protective packaging, annual requirement",
    "To: Ms. Ananya Rao, Procurement Analyst, Northwind Consumer Products Ltd, Bangalore",
    "Date: 27 August 2026     Validity: 30 days from date of issue",
    "Contact: Rajesh Menon, Head - Institutional Sales, rajesh.menon@shakticorrugators.co.in, +91 99860 41220",
    "Basis of pricing: rates below are PER KILOGRAM of finished board delivered, DDP Bangalore.",
    "Conversion to per-piece is at buyer's computation using the board weights stated in column 'Wt/pc (kg)'.",
]:
    c.drawString(20 * mm, y, t); y -= 4.6 * mm

y -= 3 * mm
c.setFont("Helvetica-Bold", 8)
cols = [20, 32, 100, 122, 143, 163, 180]
for x, t in zip(cols, ["Ln", "Item / grade offered", "Wt/pc (kg)", "Rate/kg INR", "Qty offered", "Lead", "Comply"]):
    c.drawString(x * mm, y, t)
y -= 2 * mm
c.line(20 * mm, y, W - 20 * mm, y)
y -= 4.5 * mm
c.setFont("Helvetica", 7.5)
page = 1
for ln in LINES:
    if y < 30 * mm:
        c.setFont("Helvetica-Oblique", 7)
        c.drawString(20 * mm, 20 * mm, "continued overleaf")
        c.showPage(); page += 1; header(c, page); y = H - 40 * mm
        c.setFont("Helvetica", 7.5)
    kg = ln["kgPerUnit"]
    rate_per_kg = round(p(ln, "shakti") / kg, 2)
    comply = "Yes"
    grade = ln["description"]
    if ln["lineNo"] == 9:
        comply = "Partial"
        grade = ln["description"] + " (3-colour flexo only, 4th colour not available)"
    if ln["lineNo"] == 24:
        comply = "Sub"
        grade = "Stretch film 500mm x 25 micron (substitute for 23 micron)"
    vals = [str(ln["lineNo"]), grade[:52], "%.2f" % kg, "%.2f" % rate_per_kg,
            str(ln["quantity"]), "%dd" % (14 if ln["_base"] < 50 else 21), comply]
    for x, t in zip(cols, vals):
        c.drawString(x * mm, y, t)
    y -= 4.2 * mm

if y < 70 * mm:
    c.showPage(); page += 1; header(c, page); y = H - 40 * mm
y -= 6 * mm
c.setFont("Helvetica-Bold", 9)
c.drawString(20 * mm, y, "Commercial notes"); y -= 5 * mm
c.setFont("Helvetica", 7.5)
notes = [
    "1. Rates are DDP Bangalore and include freight, packing and transit insurance. No separate freight will be billed.",
    "2. GST at 18% is extra on all lines. HSN 48191010 / 48192010 as applicable.",
    "3. Minimum order per line is 2,000 pieces except rolls and pallet boxes where MOQ is 100.",
    "4. Price break: 3% reduction on rate/kg for any line where the released quantity exceeds 20,000 pieces.",
    "5. Payment 30 days net. Early settlement within 10 days attracts 2% cash discount.",
    "6. Kraft paper is imported; rates are firm for 90 days after which they follow the RISI kraft index.",
    "7. Weekly staggered despatch is available at no extra cost. Partial delivery accepted.",
    "8. Replacement for damaged/rejected board within 72 hours of intimation.",
    "9. Local warehouse at Bommasandra holds approximately 3 weeks of buffer for repeat lines.",
]
for n in notes:
    c.drawString(20 * mm, y, n); y -= 4.2 * mm
y -= 4 * mm
c.setFont("Helvetica-Oblique", 7)
c.drawString(20 * mm, y, "* An additional across-the-board settlement discount of 4% applies to the total invoice value if the entire")
y -= 3.6 * mm
c.drawString(20 * mm, y, "  requirement of all thirty lines is placed on Shakti Corrugators as a single annual contract.")

c.showPage(); page += 1; header(c, page); y = H - 42 * mm
c.setFont("Helvetica-Bold", 10)
c.drawString(20 * mm, y, "Annexure A - Buyer questionnaire"); y -= 7 * mm
c.setFont("Helvetica", 8)
sh_ans = ["Yes - ISO 9001:2015, certificate SCL/QMS/2024, valid to Jun-2027",
          "No - FSC CoC application submitted, audit scheduled Nov 2026",
          "Approximately 1,250 MT per month across two plants",
          "Yes - NABL accredited in-house lab, bursting strength and ECT",
          "31 years", "Yes - 3 weeks buffer at Bommasandra warehouse",
          "96.2% OTIF for FY25-26 (source: our SAP despatch report)",
          "Yes - 60 day terms acceptable for annual contract"]
for (qid, q, *_), a in zip(QUESTIONNAIRE, sh_ans):
    c.setFont("Helvetica-Bold", 8); c.drawString(20 * mm, y, qid + ". " + q); y -= 4.4 * mm
    c.setFont("Helvetica", 8); c.drawString(25 * mm, y, a); y -= 6.5 * mm
c.save()

# ============================================================== 3. NIRVANA DOCX
def make_docx(path, paragraphs):
    def esc(s):
        return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    body = ""
    for style, text in paragraphs:
        b = ' <w:b/>' if style in ("h1", "h2") else ""
        sz = {"h1": "32", "h2": "26"}.get(style, "20")
        body += (
            '<w:p><w:pPr><w:rPr>%s<w:sz w:val="%s"/></w:rPr></w:pPr>'
            '<w:r><w:rPr>%s<w:sz w:val="%s"/></w:rPr><w:t xml:space="preserve">%s</w:t></w:r></w:p>'
            % (b, sz, b, sz, esc(text))
        )
    doc = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
           '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
           '<w:body>' + body + '</w:body></w:document>')
    ct = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
          '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
          '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
          '<Default Extension="xml" ContentType="application/xml"/>'
          '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
          '</Types>')
    rels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
            '</Relationships>')
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", ct)
        z.writestr("_rels/.rels", rels)
        z.writestr("word/document.xml", doc)

paras = [
    ("h1", "Nirvana Packaging Works"),
    ("p", "18/3 Peenya 2nd Stage, Bangalore 560058, Karnataka  |  GSTIN 29AAGFN1182C1ZK  |  Firm registration KA/PTN/2011/3390"),
    ("p", "Primary contact: Farah Qureshi, Partner - Sales. farah@nirvanapackaging.in  /  +91 80 2839 4410"),
    ("h2", "Subject: Our offer against RFQ-2026-0417"),
    ("p", "Dear Ms. Rao, thank you for including Nirvana Packaging Works in your enquiry for corrugated and protective packaging. We are pleased to submit our proposal below. Please note we have quoted twenty seven of the thirty lines; we have regretted lines 24 (stretch wrap film), 22 (PP strapping) and 30 (thermal labels) as these are traded items outside our manufacturing scope and we prefer not to quote what we do not make."),
    ("p", "All rates quoted are in Indian Rupees, per piece, EXCLUSIVE of GST which will be charged at 18 percent. Rates are ex-our-works Peenya. Delivery to your Bangalore stores is offered at a flat two lakh eighty thousand rupees for the full annual requirement, or at actuals if you prefer to nominate your own transporter. Insurance in transit is to buyer's account."),
    ("h2", "Rates"),
]
for ln in LINES:
    if ln["lineNo"] in (22, 24, 30):
        continue
    rate = p(ln, "nirvana")
    extra = ""
    if ln["lineNo"] == 5:
        extra = " We can offer this only in 200 GSM 26 BF, not 28 BF; please treat as a partial compliance."
    if ln["lineNo"] == 19:
        extra = " Offered as a substitute construction with a separate top cap instead of an integral lid."
    paras.append(("p",
        "Line %d, %s, %s. Our rate is Rs. %.2f per %s. Minimum order quantity %d pieces. We can deliver in %d days from firm order and we confirm availability of the full quantity of %s %s.%s"
        % (ln["lineNo"], ln["sku"], ln["description"], rate, ln["uom"],
           1000 if ln["_base"] < 50 else 250,
           15 if ln["_base"] < 50 else 25,
           f"{ln['quantity']:,}", ln["uom"], extra)))

paras += [
    ("h2", "Discounts and other commercial terms"),
    ("p", "A trade discount of three percent is applicable on all lines where the annual released quantity crosses fifteen thousand pieces. Over and above this, a volume rebate of one and a half percent of the total contract value will be issued as a credit note at the end of the contract year if the total offtake exceeds one crore twenty lakh rupees. Should you settle our invoices within fifteen days we will extend a further one percent early payment discount."),
    ("p", "This quotation is valid for thirty days from today. Prices are firm for the first six months of the contract; thereafter we would seek a revision linked to the movement in kraft paper landed cost, capped at plus or minus seven percent in any single revision."),
    ("p", "Payment terms sought are thirty days from date of invoice. We are unable to accept sixty day terms at present given our own working capital position."),
    ("p", "Partial deliveries are acceptable and we can work to a weekly despatch calendar. In the event of a back order the balance quantity is despatched within ten working days and the freight on the balance despatch is to our account. Rejected or damaged board is replaced within five working days. We do not hold buffer stock at your location but we do keep roughly one week of finished stock at Peenya for repeat lines."),
    ("h2", "Responses to your questionnaire"),
    ("p", "Q1. We hold ISO 9001:2015 certification, certificate number NPW/9001/2023, valid until August 2026 and currently under renewal audit."),
    ("p", "Q2. We do not hold FSC chain of custody certification."),
    ("p", "Q3. Our monthly corrugation capacity is about three hundred and twenty metric tonnes."),
    ("p", "Q4. We have a basic in-house lab with a bursting strength tester. Edge crush testing is outsourced to an accredited laboratory."),
    ("p", "Q5. The firm has been supplying corrugated packaging for fourteen years."),
    ("p", "Q6. We are not able to commit to two weeks of buffer stock at present."),
    ("p", "Q7. Our on time in full performance for the last twelve months was approximately eighty nine percent."),
    ("p", "Q8. As stated above, we are unable to accept sixty day payment terms."),
    ("p", "We look forward to your favourable consideration. Yours sincerely, Farah Qureshi, Partner."),
]
make_docx(os.path.join(OUT, "nirvana-packaging-quote.docx"), paras)

# ====================================================== 4. PACIFIC PACK - PHOTO
from PIL import Image, ImageDraw, ImageFont, ImageFilter

def font(sz, bold=False):
    for cand in ["/usr/share/fonts/truetype/dejavu/DejaVuSans%s.ttf" % ("-Bold" if bold else ""),
                 "/usr/share/fonts/truetype/liberation/LiberationSans%s.ttf" % ("-Bold" if bold else "")]:
        if os.path.exists(cand):
            return ImageFont.truetype(cand, sz)
    return ImageFont.load_default()

CARD_W, CARD_H = 1500, 2100
card = Image.new("RGB", (CARD_W, CARD_H), (250, 247, 238))
d = ImageDraw.Draw(card)
d.rectangle([0, 0, CARD_W, 150], fill=(24, 58, 74))
d.text((60, 40), "PACIFIC PACK GLOBAL PTE LTD", font=font(48, True), fill=(255, 255, 255))
d.text((60, 100), "18 Tuas Link 2, Singapore 638564  |  UEN 200914772K", font=font(24), fill=(200, 220, 230))
y = 190
d.text((60, y), "PRINTED RATE CARD - RFQ-2026-0417", font=font(34, True), fill=(24, 58, 74)); y += 50
d.text((60, y), "Valid 25 Aug 2026 - 24 Sep 2026.  ALL PRICES IN USD, CIF Chennai.", font=font(26), fill=(40, 40, 40)); y += 38
d.text((60, y), "Pricing unit: PER BOX OF 100 PIECES unless the line says ROLL or SHEET.", font=font(26), fill=(40, 40, 40)); y += 38
d.text((60, y), "Contact: Melissa Tan, melissa.tan@pacificpack.sg, +65 6863 1180", font=font(24), fill=(70, 70, 70)); y += 46
d.line([60, y, CARD_W - 60, y], fill=(24, 58, 74), width=3); y += 16
d.text((60, y), "LN", font=font(24, True), fill=(0, 0, 0))
d.text((140, y), "ITEM", font=font(24, True), fill=(0, 0, 0))
d.text((980, y), "UNIT", font=font(24, True), fill=(0, 0, 0))
d.text((1230, y), "USD", font=font(24, True), fill=(0, 0, 0))
y += 40
USD = 87.4
for ln in LINES[:26]:
    if ln["uom"] == "piece":
        unit, mult = "box/100", 100
    else:
        unit, mult = ln["uom"], 1
    usd = round(p(ln, "pacific") * mult / USD, 2)
    d.text((60, y), str(ln["lineNo"]), font=font(23), fill=(20, 20, 20))
    d.text((140, y), ln["description"][:44], font=font(23), fill=(20, 20, 20))
    d.text((980, y), unit, font=font(23), fill=(20, 20, 20))
    d.text((1230, y), "%,.2f".replace("%,", "%") % usd, font=font(23, True), fill=(20, 20, 20))
    y += 34
y += 12
d.line([60, y, CARD_W - 60, y], fill=(120, 120, 120), width=2); y += 18
for t in [
    "Lines 27-30 quoted on request - not stocked in Singapore.",
    "Ocean freight & insurance included (CIF). Indian customs duty, IGST 18% and",
    "inland haulage Chennai-Bangalore to buyer's account (est. USD 3,400 lump sum).",
    "Lead time 35 days ex-Singapore. MOQ 50 boxes per line. Partial shipment allowed.",
    "Payment: 30% advance, balance against B/L copy. Validity 30 days.",
    "Volume discount 4% above USD 150,000 order value. No early payment discount.",
    "ISO 9001 yes. FSC CoC yes (SGS-COC-009114). Capacity 900 MT/mo. Lab yes.",
    "22 years in business. Buffer stock: no. OTIF 91%. 60-day terms: no, 30 days only.",
]:
    d.text((60, y), t, font=font(24), fill=(35, 35, 35)); y += 36

# Now photograph it: perspective warp, uneven lighting, shadow, blur, jpeg noise
card = card.crop((0, 0, CARD_W, y + 70))
CW, CH = card.size
card = card.filter(ImageFilter.GaussianBlur(0.4))
BW, BH = CW + 260, CH + 260
bg = Image.new("RGB", (BW, BH), (58, 54, 48))
bgd = ImageDraw.Draw(bg)
for i in range(0, BH, 7):
    bgd.line([(0, i), (BW, i)], fill=(58 + (i % 5), 54 + (i % 4), 48), width=1)

src_corners = [(0, 0), (CW, 0), (CW, CH), (0, CH)]
dst = [(150, 175), (BW - 105, 95), (BW - 150, BH - 130), (95, BH - 195)]

def find_coeffs(pa, pb):
    matrix = []
    for pp1, pp2 in zip(pa, pb):
        matrix.append([pp2[0], pp2[1], 1, 0, 0, 0, -pp1[0] * pp2[0], -pp1[0] * pp2[1]])
        matrix.append([0, 0, 0, pp2[0], pp2[1], 1, -pp1[1] * pp2[0], -pp1[1] * pp2[1]])
    import numpy as np
    A = np.matrix(matrix, dtype=float)
    B = np.array(pa).reshape(8)
    res = np.dot(np.linalg.inv(A.T * A) * A.T, B)
    return np.array(res).reshape(8)

coeffs = find_coeffs(src_corners, dst)
warped = card.transform((BW, BH), Image.PERSPECTIVE, coeffs, Image.BICUBIC)
mask = Image.new("L", (CW, CH), 255).transform((BW, BH), Image.PERSPECTIVE, coeffs, Image.BICUBIC)
bg.paste(warped, (0, 0), mask)

light = Image.new("L", (BW, BH), 0)
ld = ImageDraw.Draw(light)
for r in range(900, 0, -20):
    ld.ellipse([420 - r * 1.4, 300 - r, 420 + r * 1.4, 300 + r], fill=int(60 * (1 - r / 900)))
light = light.filter(ImageFilter.GaussianBlur(160))
bg = Image.composite(Image.new("RGB", bg.size, (255, 255, 250)), bg, light.point(lambda v: int(v * 0.5)))
shade = Image.new("L", (BW, BH), 0)
sd = ImageDraw.Draw(shade)
sd.polygon([(BW - 420, 0), (BW, 0), (BW, BH), (BW - 700, BH)], fill=70)
shade = shade.filter(ImageFilter.GaussianBlur(220))
bg = Image.composite(Image.new("RGB", bg.size, (20, 18, 16)), bg, shade)
bg = bg.rotate(-1.8, resample=Image.BICUBIC, expand=False, fillcolor=(58, 54, 48))
bg = bg.filter(ImageFilter.GaussianBlur(0.8))
scale = 1500 / BW
bg = bg.resize((1500, int(BH * scale)), Image.LANCZOS)
bg.save(os.path.join(OUT, "pacific-pack-ratecard-photo.jpg"), quality=64)

# =============================================================== 5. VINDHYA TXT
vind = """From: prakash@vindhyaboxes.com
To: ananya.rao@northwindcp.in
Subject: RE: RFQ-2026-0417 - Corrugated packaging annual requirement
Date: Fri, 28 Aug 2026 19:42 +0530

Ananya madam,

Rates as discussed. Rs 42/kg for the 5-ply and 7-ply items, 38 for the 3-ply and
2-ply, rest same as last year. Freight extra, about 2.6 lakhs for the year to your
Bangalore godown, we will bill at actuals. GST 18% extra as usual.

For the non-board items - edge protector Rs 13.90 a piece, PP strap roll Rs 720,
BOPP printed tape Rs 36.50 a roll, stretch film Rs 1385 a roll, honeycomb pad
Rs 112 each, thermal labels Rs 205 a roll, EPE liner Rs 33 each. Pallet box and
sleeve we will do at Rs 615 and Rs 372 respectively.

Board weights are the same as what we supplied you in FY25 so you can work out the
per piece rate from the kg rate directly.

Lead time 10-12 days for board, 3 weeks for the printed tape. We can do weekly
despatch. Payment 30 days please, we cannot stretch to 60. Validity 3 weeks.
Discount 2% if the whole thing comes to us.

Questionnaire - ISO 9001 yes (certificate with our Bangalore office, I will send
separately). FSC no. Capacity around 180 MT a month. Testing lab - we have a
bursting strength tester only. In the business 26 years. Buffer stock yes we can
keep 2 weeks. OTIF last year I would say around 87-88%. 60 day payment no.

Anything else needed please call.

Prakash Nayak
Vindhya Boxes & Cartons
No 7, Yeshwanthpur Industrial Suburb, Bangalore 560022
GSTIN 29AACFV3391D1ZP | +91 98801 33427
"""
open(os.path.join(OUT, "vindhya-boxes-email.txt"), "w").write(vind)

# ======================================================== rfq.ts (typed data)
for ln in LINES:
    ln.pop("_base")

rfq = {
    "id": "RFQ-2026-0417",
    "title": "Annual Requirement - Corrugated & Protective Packaging, Bangalore Plant",
    "buyingOrganization": "Northwind Consumer Products Ltd",
    "businessUnit": "Supply Chain - Packaging",
    "buyerContact": "Ananya Rao",
    "buyerEmail": "ananya.rao@northwindcp.in",
    "productCategory": "Packaging - Corrugated & protective",
    "purpose": "Replace three expiring regional contracts with a single annual rate contract",
    "issueDate": str(ISSUE),
    "questionsDeadline": str(ISSUE + timedelta(days=10)),
    "submissionDeadline": str(ISSUE + timedelta(days=15)),
    "expectedAwardDate": str(ISSUE + timedelta(days=30)),
    "quoteValidity": "1 month",
    "deliveryLocation": "Bangalore, India",
    "expectedDeliveryDate": str(ISSUE + timedelta(days=60)),
    "contractDuration": "12 months",
    "estimatedTotalQuantity": sum(l["quantity"] for l in LINES),
    "currency": "INR",
    "taxable": True,
    "partialQuotationAllowed": True,
    "alternativeProductAllowed": True,
    "partialAwardAllowed": True,
    "submissionInstructions": "Reply to this email with your quote and the completed questionnaire attached. Quote RFQ-2026-0417 in the subject line. Any document format is accepted.",
    "lineItems": LINES,
    "questionnaire": [
        {"id": q[0], "question": q[1], "type": q[2], "target": q[3], "weight": q[4]}
        for q in QUESTIONNAIRE
    ],
}

ts = "// AUTO-GENERATED by scripts/gen_fixtures.py -- do not edit by hand.\n"
ts += "import type { Rfq } from \"@/lib/types\";\n\n"
ts += "export const rfq: Rfq = " + json.dumps(rfq, indent=2) + " as const;\n"
open(os.path.join(ROOT, "src", "data", "rfq.ts"), "w").write(ts)

print("wrote", os.listdir(OUT))
